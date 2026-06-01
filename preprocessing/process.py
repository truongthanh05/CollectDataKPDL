import re
import unicodedata
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


# Ẩn cảnh báo PerformanceWarning, không ảnh hưởng kết quả
warnings.simplefilter(
    action="ignore",
    category=pd.errors.PerformanceWarning
)


# =====================================================
# CẤU HÌNH FILE
# =====================================================

# Cấu trúc:
# tourism_data_project/
# ├── data/
# │   ├── raw/
# │   │   └── traveloka_hotels_full_23_Copy.xlsx
# │   └── processed/
# └── preprocessing/
#     └── process.py

BASE_DIR = Path(__file__).resolve().parents[1]

INPUT_FILE = BASE_DIR / "data" / "raw" / "traveloka_hotels_HaNoi.xlsx"

OUTPUT_DIR = BASE_DIR / "data" / "processed"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CLEANED_FILE = OUTPUT_DIR / "traveloka_hotels_cleaned.xlsx"
TRANSACTION_FILE = OUTPUT_DIR / "traveloka_transactions.csv"
ONEHOT_FILE = OUTPUT_DIR / "traveloka_apriori_onehot.csv"
REPORT_FILE = OUTPUT_DIR / "traveloka_processed_report.xlsx"


# =====================================================
# HÀM PHỤ TRỢ
# =====================================================

def normalize_space(text):
    """Xóa khoảng trắng dư."""
    if pd.isna(text):
        return ""

    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)

    return text


def remove_accents(text):
    """Bỏ dấu tiếng Việt để tạo item cho Apriori."""
    text = str(text)

    text = unicodedata.normalize("NFD", text)

    text = "".join(
        char for char in text
        if unicodedata.category(char) != "Mn"
    )

    return text


def make_item(prefix, value):
    """
    Chuyển text thành item dạng:
    'Gần sân bay' -> 'nearby_gan_san_bay'
    """
    value = normalize_space(value).lower()
    value = remove_accents(value)

    value = re.sub(r"[^a-zA-Z0-9]+", "_", value)
    value = value.strip("_")

    if value == "":
        return ""

    return f"{prefix}_{value}"


def unique_keep_order(items):
    """Xóa trùng nhưng giữ thứ tự."""
    result = []
    seen = set()

    for item in items:
        item = normalize_space(item)

        if item and item not in seen:
            result.append(item)
            seen.add(item)

    return result


def split_items(value, split_comma=False):
    """
    Tách dữ liệu trong 1 ô.

    Ví dụ:
    'Wifi | Máy lạnh | Thang máy'
    -> ['Wifi', 'Máy lạnh', 'Thang máy']

    Nếu split_comma=True:
    'Máy lạnh, Lễ tân 24h, WiFi'
    -> ['Máy lạnh', 'Lễ tân 24h', 'WiFi']
    """
    if pd.isna(value):
        return []

    text = str(value)
    text = text.replace("\n", " | ")

    if split_comma:
        parts = re.split(r"\s*\|\s*|\s*,\s*|\s*;\s*", text)
    else:
        parts = re.split(r"\s*\|\s*|\s*;\s*", text)

    parts = [
        normalize_space(part)
        for part in parts
        if normalize_space(part) != ""
    ]

    return unique_keep_order(parts)


def clean_price(value):
    """
    '590.586 VND' -> 590586
    """
    if pd.isna(value):
        return None

    text = str(value)

    digits = re.sub(r"[^\d]", "", text)

    if digits == "":
        return None

    return int(digits)


def clean_float(value):
    """
    '8,9' -> 8.9
    '8.9' -> 8.9
    """
    if pd.isna(value):
        return None

    text = str(value).strip()
    text = text.replace(",", ".")

    match = re.search(r"\d+(\.\d+)?", text)

    if not match:
        return None

    return float(match.group())


def clean_int(value):
    """
    'Từ 36 đánh giá' -> 36
    '36' -> 36
    """
    if pd.isna(value):
        return None

    text = str(value)

    match = re.search(r"\d+", text)

    if not match:
        return None

    return int(match.group())


def extract_district(address):
    """
    Tách quận/huyện từ địa chỉ.
    """
    if pd.isna(address):
        return ""

    text = str(address)

    patterns = [
        r"(Quận\s*[^\.,]+)",
        r"(Huyện\s*[^\.,]+)",
        r"(Thành phố Thủ Đức)",
        r"(TP\.?\s*Thủ Đức)",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE
        )

        if match:
            return normalize_space(match.group(1))

    return ""


def price_level(price):
    """
    Phân mức giá để đưa vào luật kết hợp.
    """
    if pd.isna(price):
        return ""

    if price < 500000:
        return "Giá rẻ"

    if price < 1000000:
        return "Giá trung bình"

    return "Giá cao"


def rating_level(rating):
    """
    Phân mức rating.
    """
    if pd.isna(rating):
        return ""

    if rating >= 9:
        return "Rating xuất sắc"

    if rating >= 8:
        return "Rating cao"

    if rating >= 7:
        return "Rating khá"

    return "Rating thấp"


def clean_nearby_item(item):
    """
    'Sân bay Tân Sơn Nhất (1.2 km)' -> 'Sân bay Tân Sơn Nhất'
    """
    item = normalize_space(item)

    item = re.sub(r"\([^)]*\)", "", item)

    return normalize_space(item)


def infer_tourism_tags(nearby_text):
    """
    Tạo tag du lịch từ địa điểm gần khách sạn.
    """
    text = normalize_space(nearby_text).lower()

    tags = []

    if any(k in text for k in ["sân bay", "airport", "tân sơn nhất"]):
        tags.append("Gần sân bay / transit")

    if any(k in text for k in ["chợ", "market", "mall", "plaza", "centre", "center", "vincom", "takashimaya"]):
        tags.append("Mua sắm / ăn uống")

    if any(k in text for k in ["công viên", "park", "thảo cầm viên"]):
        tags.append("Gia đình / thư giãn")

    if any(k in text for k in ["bệnh viện", "hospital", "phòng khám"]):
        tags.append("Lưu trú y tế")

    if any(k in text for k in ["sân vận động", "stadium", "quân khu"]):
        tags.append("Sự kiện / giải trí")

    if any(k in text for k in ["bảo tàng", "nhà thờ", "dinh", "phố đi bộ", "nguyễn huệ", "bùi viện", "bitexco"]):
        tags.append("Tham quan thành phố")

    return " | ".join(
        unique_keep_order(tags)
    )


def merge_dynamic_columns(df, keyword, new_col, split_comma=False):
    """
    Gộp các cột có tên chứa keyword thành 1 cột duy nhất.

    Ví dụ:
    general_Những tiện nghi khác tại Hotel Nikko Saigon
    general_Những tiện nghi khác tại Fusion Original Saigon Centre

    -> general_Những tiện nghi khác
    """
    cols = [
        col for col in df.columns
        if keyword in col
    ]

    if not cols:
        df[new_col] = ""
        return df

    def merge_row(row):
        items = []

        for col in cols:
            items.extend(
                split_items(
                    row.get(col, ""),
                    split_comma=split_comma
                )
            )

        return " | ".join(
            unique_keep_order(items)
        )

    df[new_col] = df.apply(
        merge_row,
        axis=1
    )

    df = df.drop(
        columns=cols,
        errors="ignore"
    )

    return df.copy()


def set_excel_style(writer, sheet_df_map):
    """
    Format Excel cho dễ nhìn.
    Dùng openpyxl.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    normal_alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    for sheet_name, data_df in sheet_df_map.items():

        worksheet = writer.sheets[sheet_name]

        # Freeze dòng tiêu đề
        worksheet.freeze_panes = "A2"

        # Bật filter
        if len(data_df.columns) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        # Style header
        for cell in worksheet[1]:

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set độ rộng cột
        for col_idx, col_name in enumerate(
            data_df.columns,
            start=1
        ):

            col_letter = worksheet.cell(
                row=1,
                column=col_idx
            ).column_letter

            col_name_str = str(col_name)

            if col_name_str in [
                "hotel_name",
                "hotel_url",
                "address",
                "nearby_places",
                "nearby_places_clean",
                "facilities_clean",
                "transaction_items",
            ]:
                worksheet.column_dimensions[col_letter].width = 45

            elif col_name_str.startswith("facility_"):
                worksheet.column_dimensions[col_letter].width = 35

            elif col_name_str.startswith("general_"):
                worksheet.column_dimensions[col_letter].width = 35

            elif col_name_str.startswith("score_"):
                worksheet.column_dimensions[col_letter].width = 16

            elif col_name_str in [
                "price_clean",
                "review_count_clean",
                "star_rating_clean",
                "overall_rating_clean",
            ]:
                worksheet.column_dimensions[col_letter].width = 18

            else:
                worksheet.column_dimensions[col_letter].width = 22

        # Wrap text toàn bộ sheet
        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = normal_alignment


# =====================================================
# ĐỌC FILE
# =====================================================

df = pd.read_excel(INPUT_FILE)

original_col_count = len(df.columns)

df.columns = [
    normalize_space(col)
    for col in df.columns
]

print("Số dòng ban đầu:", len(df))
print("Số cột ban đầu:", len(df.columns))


# =====================================================
# 1. XÓA DÒNG TRỐNG + TRÙNG
# =====================================================

df = df.dropna(how="all")

if "hotel_url" in df.columns:
    df = df.drop_duplicates(
        subset=["hotel_url"]
    )
else:
    df = df.drop_duplicates()


# =====================================================
# 2. GỘP CÁC CỘT DYNAMIC KHÔNG HỢP LÝ
# =====================================================

df = merge_dynamic_columns(
    df,
    keyword="general_Những tiện nghi khác tại",
    new_col="general_Những tiện nghi khác",
    split_comma=True
)

df = merge_dynamic_columns(
    df,
    keyword="general_Những địa điểm thú vị gần đó",
    new_col="general_Những địa điểm thú vị gần đó",
    split_comma=True
)


# =====================================================
# 3. XÓA CÁC CỘT KHÔNG HỢP LÝ
# =====================================================

drop_keywords = [
    "general_Số phòng còn trống tai",
    "general_Số phòng còn trống tại",
    "general_Số lầu tại",
]

drop_cols = []

for col in df.columns:

    for keyword in drop_keywords:

        if keyword in col:
            drop_cols.append(col)

drop_cols = list(set(drop_cols))

df = df.drop(
    columns=drop_cols,
    errors="ignore"
)

df = df.copy()

print("Đã xóa số cột rác:", len(drop_cols))


# =====================================================
# 4. LÀM SẠCH CÁC CỘT SỐ
# =====================================================

if "price" in df.columns:
    df["price_clean"] = df["price"].apply(clean_price)

if "star_rating" in df.columns:
    df["star_rating_clean"] = df["star_rating"].apply(clean_float)

if "overall_rating" in df.columns:
    df["overall_rating_clean"] = df["overall_rating"].apply(clean_float)

if "review_count" in df.columns:
    df["review_count_clean"] = df["review_count"].apply(clean_int)

# =====================================================
# 4.1. XÓA DÒNG THIẾU PRICE HOẶC OVERALL_RATING
# =====================================================

before_drop_required = len(df)

required_cols = []

if "price_clean" in df.columns:
    required_cols.append("price_clean")

if "overall_rating_clean" in df.columns:
    required_cols.append("overall_rating_clean")

if required_cols:
    df = df.dropna(
        subset=required_cols
    ).copy()

removed_missing_price_rating = before_drop_required - len(df)

print(
    "Đã xóa số dòng thiếu price hoặc overall_rating:",
    removed_missing_price_rating
)
score_cols = [
    col for col in df.columns
    if col.startswith("score_")
]

for col in score_cols:
    df[col + "_clean"] = df[col].apply(clean_float)


# =====================================================
# 5. TÁCH ĐỊA CHỈ
# =====================================================

if "address" in df.columns:
    df["district"] = df["address"].apply(extract_district)
else:
    df["district"] = ""


# =====================================================
# 6. XỬ LÝ ĐỊA ĐIỂM NỔI BẬT
# =====================================================

nearby_source_cols = []

for col in [
    "nearby_places",
    "general_Điểm đến phổ biến",
    "general_Những địa điểm thú vị gần đó",
]:

    if col in df.columns:
        nearby_source_cols.append(col)


def build_nearby_places(row):
    items = []

    for col in nearby_source_cols:

        raw_items = split_items(
            row.get(col, ""),
            split_comma=False
        )

        for item in raw_items:
            items.append(
                clean_nearby_item(item)
            )

    return " | ".join(
        unique_keep_order(items)
    )


df["nearby_places_clean"] = df.apply(
    build_nearby_places,
    axis=1
)

df["tourism_tags"] = df["nearby_places_clean"].apply(
    infer_tourism_tags
)


# =====================================================
# 7. XỬ LÝ TIỆN ÍCH
# =====================================================

facility_cols = [
    col for col in df.columns
    if col.startswith("facility_")
]

extra_facility_cols = []

for col in [
    "general_Tiện ích chung",
    "general_Những tiện nghi khác",
]:

    if col in df.columns:
        extra_facility_cols.append(col)


def build_facilities(row):
    items = []

    for col in facility_cols:

        items.extend(
            split_items(
                row.get(col, ""),
                split_comma=False
            )
        )

    for col in extra_facility_cols:

        items.extend(
            split_items(
                row.get(col, ""),
                split_comma=True
            )
        )

    return " | ".join(
        unique_keep_order(items)
    )


df["facilities_clean"] = df.apply(
    build_facilities,
    axis=1
)


# =====================================================
# 8. TẠO CỘT PHÂN NHÓM
# =====================================================

if "price_clean" in df.columns:
    df["price_level"] = df["price_clean"].apply(price_level)
else:
    df["price_level"] = ""

if "overall_rating_clean" in df.columns:
    df["rating_level"] = df["overall_rating_clean"].apply(rating_level)
else:
    df["rating_level"] = ""


# =====================================================
# 9. TẠO TRANSACTION CHO APRIORI
# =====================================================

def build_transaction(row):
    items = []

    hotel_type_cols = [
        "price_level",
        "rating_level",
        "district",
        "tourism_tags",
    ]

    for col in hotel_type_cols:

        values = split_items(
            row.get(col, ""),
            split_comma=False
        )

        for value in values:

            prefix = col.replace("_level", "")
            item = make_item(
                prefix,
                value
            )

            if item:
                items.append(item)

    if not pd.isna(
        row.get("star_rating_clean", None)
    ):
        star = int(row["star_rating_clean"])
        items.append(f"star_{star}_sao")

    facility_values = split_items(
        row.get("facilities_clean", ""),
        split_comma=False
    )

    for facility in facility_values:

        item = make_item(
            "facility",
            facility
        )

        if item:
            items.append(item)

    tourism_values = split_items(
        row.get("tourism_tags", ""),
        split_comma=False
    )

    for tag in tourism_values:

        item = make_item(
            "tourism",
            tag
        )

        if item:
            items.append(item)

    return " | ".join(
        unique_keep_order(items)
    )


df["transaction_items"] = df.apply(
    build_transaction,
    axis=1
)


# =====================================================
# 10. TẠO FILE ONE-HOT CHO APRIORI
# =====================================================

transactions = df["transaction_items"].apply(
    lambda x: split_items(
        x,
        split_comma=False
    )
)

all_items = sorted(
    set(
        item
        for trans in transactions
        for item in trans
    )
)

onehot_rows = []

for trans in transactions:

    trans_set = set(trans)

    row = {
        item: int(item in trans_set)
        for item in all_items
    }

    onehot_rows.append(row)

onehot_df = pd.DataFrame(onehot_rows)

if "hotel_name" in df.columns:
    onehot_df.insert(
        0,
        "hotel_name",
        df["hotel_name"]
    )


# =====================================================
# 11. TẠO CÁC SHEET DỄ NHÌN
# =====================================================

nearby_rows = []

for _, row in df.iterrows():

    hotel_name = row.get("hotel_name", "")

    nearby_list = split_items(
        row.get("nearby_places_clean", ""),
        split_comma=False
    )

    for place in nearby_list:

        nearby_rows.append({
            "hotel_name": hotel_name,
            "nearby_place": place
        })

nearby_df = pd.DataFrame(
    nearby_rows,
    columns=[
        "hotel_name",
        "nearby_place"
    ]
)


facility_rows = []

for _, row in df.iterrows():

    hotel_name = row.get("hotel_name", "")

    facility_list = split_items(
        row.get("facilities_clean", ""),
        split_comma=False
    )

    for facility in facility_list:

        facility_rows.append({
            "hotel_name": hotel_name,
            "facility": facility
        })

facility_df = pd.DataFrame(
    facility_rows,
    columns=[
        "hotel_name",
        "facility"
    ]
)


transaction_df = df[
    [
        "hotel_name",
        "hotel_url",
        "transaction_items"
    ]
].copy()


summary_df = pd.DataFrame({
    "Chỉ số": [
        "Số khách sạn",
        "Số cột ban đầu",
        "Số cột sau xử lý",
        "Số cột rác đã xóa",
        "Số địa điểm nearby đã tách",
        "Số tiện ích đã tách",
        "Số item Apriori",
    ],
    "Giá trị": [
        len(df),
        original_col_count,
        len(df.columns),
        len(drop_cols),
        len(nearby_df),
        len(facility_df),
        len(onehot_df.columns) - 1
        if "hotel_name" in onehot_df.columns
        else len(onehot_df.columns),
    ]
})


# =====================================================
# 12. LƯU FILE
# =====================================================

df.to_excel(
    CLEANED_FILE,
    index=False
)

transaction_df.to_csv(
    TRANSACTION_FILE,
    index=False,
    encoding="utf-8-sig"
)

onehot_df.to_csv(
    ONEHOT_FILE,
    index=False,
    encoding="utf-8-sig"
)

sheet_df_map = {
    "summary": summary_df,
    "cleaned_data": df,
    "nearby_places": nearby_df,
    "facilities": facility_df,
    "transactions": transaction_df,
    "apriori_onehot": onehot_df,
}

with pd.ExcelWriter(
    REPORT_FILE,
    engine="openpyxl"
) as writer:

    for sheet_name, sheet_df in sheet_df_map.items():

        sheet_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )

    set_excel_style(
        writer,
        sheet_df_map
    )


print("Số dòng sau xử lý:", len(df))
print("Số cột sau xử lý:", len(df.columns))

print("Đã lưu file cleaned:", CLEANED_FILE)
print("Đã lưu file transaction:", TRANSACTION_FILE)
print("Đã lưu file one-hot:", ONEHOT_FILE)
print("Đã lưu file Excel report:", REPORT_FILE)