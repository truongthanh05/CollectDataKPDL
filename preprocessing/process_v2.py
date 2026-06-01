import re
import unicodedata
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


# Ẩn cảnh báo PerformanceWarning, không ảnh hưởng kết quả
warnings.simplefilter(
    action="ignore",
    category=pd.errors.PerformanceWarning
)


# =====================================================
# CẤU HÌNH FILE
# =====================================================

# Cấu trúc:
# tourism_data_project/
# ├── data/
# │   ├── raw/
# │   │   └── traveloka_hotels_HaNoi.xlsx
# │   └── processed/
# └── preprocessing/
#     └── process.py

BASE_DIR = Path(__file__).resolve().parents[1]

INPUT_FILE = BASE_DIR / "data" / "raw" / "traveloka_hotels_HaNoi.xlsx"

OUTPUT_DIR = BASE_DIR / "data" / "processed"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CLEANED_FILE = OUTPUT_DIR / "traveloka_hotels_cleaned.xlsx"
TRANSACTION_FILE = OUTPUT_DIR / "traveloka_transactions.csv"
ONEHOT_FILE = OUTPUT_DIR / "traveloka_apriori_onehot.csv"
REPORT_FILE = OUTPUT_DIR / "traveloka_processed_report.xlsx"


# =====================================================
# HÀM PHỤ TRỢ CƠ BẢN
# =====================================================

def normalize_space(text):
    """Xóa khoảng trắng dư."""
    if pd.isna(text):
        return ""

    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)

    return text


def remove_accents(text):
    """Bỏ dấu tiếng Việt để tạo item cho Apriori."""
    text = str(text)
    text = unicodedata.normalize("NFD", text)

    text = "".join(
        char for char in text
        if unicodedata.category(char) != "Mn"
    )

    return text


def text_key(text):
    """
    Chuẩn hóa text để so khớp từ khóa:
    - lower
    - bỏ dấu
    - xóa khoảng trắng dư
    """
    text = normalize_space(text).lower()
    text = remove_accents(text)
    return text


def make_item(prefix, value):
    """
    Chuyển text thành item dạng:
    'Gần sân bay' -> 'nearby_gan_san_bay'
    """
    value = normalize_space(value).lower()
    value = remove_accents(value)

    value = re.sub(r"[^a-zA-Z0-9]+", "_", value)
    value = value.strip("_")

    if value == "":
        return ""

    return f"{prefix}_{value}"


def unique_keep_order(items):
    """Xóa trùng nhưng giữ thứ tự."""
    result = []
    seen = set()

    for item in items:
        item = normalize_space(item)

        if item and item not in seen:
            result.append(item)
            seen.add(item)

    return result


def split_items(value, split_comma=False):
    """
    Tách dữ liệu trong 1 ô.

    Ví dụ:
    'Wifi | Máy lạnh | Thang máy'
    -> ['Wifi', 'Máy lạnh', 'Thang máy']

    Nếu split_comma=True:
    'Máy lạnh, Lễ tân 24h, WiFi'
    -> ['Máy lạnh', 'Lễ tân 24h', 'WiFi']
    """
    if pd.isna(value):
        return []

    text = str(value)
    text = text.replace("\n", " | ")

    if split_comma:
        parts = re.split(r"\s*\|\s*|\s*,\s*|\s*;\s*", text)
    else:
        parts = re.split(r"\s*\|\s*|\s*;\s*", text)

    parts = [
        normalize_space(part)
        for part in parts
        if normalize_space(part) != ""
    ]

    return unique_keep_order(parts)


# =====================================================
# HÀM LÀM SẠCH SỐ
# =====================================================

def clean_price(value):
    """
    '590.586 VND' -> 590586
    """
    if pd.isna(value):
        return None

    text = str(value)
    digits = re.sub(r"[^\d]", "", text)

    if digits == "":
        return None

    return int(digits)


def clean_float(value):
    """
    '8,9' -> 8.9
    '8.9' -> 8.9
    """
    if pd.isna(value):
        return None

    text = str(value).strip()
    text = text.replace(",", ".")

    match = re.search(r"\d+(\.\d+)?", text)

    if not match:
        return None

    return float(match.group())


def clean_int(value):
    """
    'Từ 36 đánh giá' -> 36
    '36' -> 36
    """
    if pd.isna(value):
        return None

    text = str(value)
    match = re.search(r"\d+", text)

    if not match:
        return None

    return int(match.group())


# =====================================================
# HÀM TÁCH ĐỊA CHỈ / KHU VỰC
# =====================================================

def extract_district(address):
    """
    Tách quận/huyện/thành phố từ địa chỉ.
    Dùng được cho nhiều vùng, không riêng TP.HCM.
    """
    if pd.isna(address):
        return ""

    text = str(address)

    patterns = [
        r"(Quận\s*[^\.,]+)",
        r"(Huyện\s*[^\.,]+)",
        r"(Thị xã\s*[^\.,]+)",
        r"(Thành phố Thủ Đức)",
        r"(TP\.?\s*Thủ Đức)",
        r"(Thành phố\s*[^\.,]+)",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE
        )

        if match:
            return normalize_space(match.group(1))

    return ""


# =====================================================
# HÀM PHÂN NHÓM GIÁ / RATING
# =====================================================

def price_level(price):
    """
    Phân mức giá để đưa vào luật kết hợp.
    """
    if pd.isna(price):
        return ""

    if price < 500000:
        return "Giá rẻ"

    if price < 1000000:
        return "Giá trung bình"

    return "Giá cao"


def rating_level(rating):
    """
    Phân mức rating.
    """
    if pd.isna(rating):
        return ""

    if rating >= 9:
        return "Rating xuất sắc"

    if rating >= 8:
        return "Rating cao"

    if rating >= 7:
        return "Rating khá"

    return "Rating thấp"


# =====================================================
# LUẬT PHÂN LOẠI ĐỊA ĐIỂM ĐA VÙNG
# =====================================================

PLACE_TYPE_RULES = {
    "airport": [
        "san bay", "airport", "tan son nhat", "noi bai", "phu bai",
        "cam ranh", "lien khuong", "da nang airport", "cat bi"
    ],

    "shopping_food": [
        "cho", "market", "mall", "plaza", "centre", "center",
        "vincom", "takashimaya", "saigon centre", "lotte",
        "coopmart", "big c", "go!", "aeon", "sieu thi",
        "pho am thuc", "food", "night market", "cho dem"
    ],

    "beach_relax": [
        "bien", "beach", "mui ne", "hon rom", "bai da", "bai sau",
        "bai truoc", "bai dai", "bai tam", "dao", "hon", "vinpearl"
    ],

    "heritage_culture": [
        "dai noi", "kinh thanh", "lang", "lang vua", "chua", "nha tho",
        "bao tang", "dinh", "thap", "pho co", "di tich", "den",
        "van mieu", "hoang thanh", "cau truong tien", "song huong",
        "thap ba", "ponagar", "mosque"
    ],

    "city_sightseeing": [
        "pho di bo", "nguyen hue", "bui vien", "bitexco",
        "cong vien", "park", "quang truong", "ho xuan huong",
        "ho hoan kiem", "ho tay", "pho duong tau", "thao cam vien",
        "garden", "cau", "ho", "old quarter", "khu pho"
    ],

    "nature_experience": [
        "doi cat", "suoi tien", "thac", "nui", "ho", "thung lung",
        "langbiang", "deo", "vuon hoa", "rung", "bau trang",
        "hon rom", "hang", "cave", "waterfall"
    ],

    "medical": [
        "benh vien", "hospital", "phong kham", "y khoa", "medical",
        "da khoa", "tam anh", "cho ray", "hoa hao"
    ],

    "event_entertainment": [
        "san van dong", "stadium", "quan khu", "nha hat",
        "rap", "cinema", "arena", "hoi nghi", "tiec cuoi",
        "theatre", "opera", "trung tam hoi nghi"
    ],

    "transport": [
        "ben xe", "ga", "station", "cang", "port",
        "bus", "tau", "metro", "ga metro", "bến xe"
    ],

    "education_office": [
        "truong", "dai hoc", "university", "cong ty", "ubnd",
        "toa nha", "van phong", "office", "cp", "corporation"
    ],
}


REGION_PLACE_TYPE_RULES = {
    "ho chi minh": {
        "airport": ["tan son nhat"],
        "shopping_food": [
            "cho ben thanh", "cho dem ben thanh", "takashimaya",
            "saigon centre", "vincom", "bui vien", "cho an dong"
        ],
        "city_sightseeing": [
            "pho di bo nguyen hue", "bitexco", "dinh doc lap",
            "nha tho duc ba", "thao cam vien", "ga metro ben thanh"
        ],
        "event_entertainment": [
            "san van dong quan khu 7", "san van dong thong nhat"
        ],
        "medical": [
            "tam anh", "cho ray", "hoa hao"
        ]
    },

    "ha noi": {
        "airport": ["noi bai"],
        "heritage_culture": [
            "van mieu", "hoang thanh thang long", "den ngoc son",
            "nha tho lon ha noi", "den bach ma", "pho co ha noi"
        ],
        "city_sightseeing": [
            "ho hoan kiem", "ho tay", "pho duong tau ha noi",
            "pho co", "ta hien"
        ],
        "shopping_food": [
            "cho dong xuan", "melinh plaza", "trang tien plaza"
        ]
    },

    "hue": {
        "heritage_culture": [
            "dai noi", "kinh thanh hue", "lang khai dinh",
            "lang tu duc", "chua thien mu", "cau truong tien",
            "song huong", "cho dong ba"
        ],
        "airport": ["phu bai"],
    },

    "phan thiet": {
        "beach_relax": [
            "mui ne", "hon rom", "bai da ong dia", "bien mui ne",
            "lang chai mui ne", "bai bien"
        ],
        "nature_experience": [
            "doi cat bay", "suoi tien", "bau trang"
        ]
    },

    "da lat": {
        "nature_experience": [
            "ho xuan huong", "thung lung tinh yeu", "langbiang",
            "thac datanla", "vuon hoa da lat"
        ],
        "shopping_food": [
            "cho da lat", "cho dem da lat"
        ]
    },

    "nha trang": {
        "beach_relax": [
            "bien nha trang", "vinwonders", "hon chong",
            "hon mun", "bai dai"
        ],
        "heritage_culture": [
            "thap ba ponagar"
        ]
    },

    "da nang": {
        "beach_relax": [
            "bien my khe", "bai bien my khe", "non nuoc"
        ],
        "city_sightseeing": [
            "cau rong", "cau song han", "cau tinh yeu"
        ],
        "heritage_culture": [
            "ngu hanh son", "bao tang cham"
        ]
    },

    "hoi an": {
        "heritage_culture": [
            "pho co hoi an", "chua cau", "nha co", "hoi quan"
        ],
        "beach_relax": [
            "bien an bang", "cua dai"
        ]
    }
}


def classify_place_type(place_name, region=""):
    """
    Phân loại địa điểm theo nhóm chuẩn hóa.
    Dùng được cho nhiều vùng: HCM, Hà Nội, Huế, Phan Thiết, Đà Lạt...
    """
    place = text_key(place_name)
    region_key = text_key(region)

    # 1. Ưu tiên luật riêng theo region
    for region_name, region_rules in REGION_PLACE_TYPE_RULES.items():

        if region_name in region_key:

            for place_type, keywords in region_rules.items():

                for keyword in keywords:

                    if text_key(keyword) in place:
                        return place_type

    # 2. Nếu không khớp region riêng thì dùng luật chung
    for place_type, keywords in PLACE_TYPE_RULES.items():

        for keyword in keywords:

            if text_key(keyword) in place:
                return place_type

    return "other"


# =====================================================
# HÀM XỬ LÝ KHOẢNG CÁCH / NEARBY
# =====================================================

def parse_distance_to_meters(distance_text):
    """
    '470 m' -> 470
    '1.71 km' -> 1710
    """
    if pd.isna(distance_text):
        return None

    text = str(distance_text).lower().strip()
    text = text.replace(",", ".")

    match = re.search(r"(\d+(\.\d+)?)", text)

    if not match:
        return None

    value = float(match.group(1))

    if "km" in text:
        return int(value * 1000)

    if "m" in text:
        return int(value)

    return None


def parse_nearby_item(item):
    """
    'Sân bay Tân Sơn Nhất (1.71 km)'
    -> place_name, distance_text, distance_m

    Nếu item không có khoảng cách, vẫn giữ place_name và distance rỗng.
    """
    item = normalize_space(item)

    distance_text = ""

    match = re.search(r"\(([^)]*)\)", item)

    if match:
        distance_text = normalize_space(match.group(1))

    place_name = re.sub(r"\([^)]*\)", "", item)
    place_name = normalize_space(place_name)

    distance_m = parse_distance_to_meters(distance_text)

    return place_name, distance_text, distance_m


def distance_level(distance_m):
    """
    Phân mức khoảng cách để dùng cho web.
    """
    if pd.isna(distance_m):
        return "Không rõ khoảng cách"

    if distance_m <= 500:
        return "Đi bộ rất gần"

    if distance_m <= 1000:
        return "Đi bộ được"

    if distance_m <= 3000:
        return "Di chuyển ngắn"

    return "Hơi xa"


def tourism_tags_from_place_types(place_types):
    """
    Tạo tourism_tags từ nhóm địa điểm.
    """
    tags = []

    if "airport" in place_types:
        tags.append("Transit / công tác gần sân bay")

    if "shopping_food" in place_types:
        tags.append("Mua sắm / ăn uống")

    if "beach_relax" in place_types:
        tags.append("Du lịch biển / nghỉ dưỡng")

    if "heritage_culture" in place_types:
        tags.append("Tham quan văn hóa / di tích")

    if "city_sightseeing" in place_types:
        tags.append("Khám phá thành phố")

    if "nature_experience" in place_types:
        tags.append("Thiên nhiên / trải nghiệm")

    if "medical" in place_types:
        tags.append("Lưu trú y tế")

    if "event_entertainment" in place_types:
        tags.append("Sự kiện / giải trí")

    if "transport" in place_types:
        tags.append("Gần điểm giao thông")

    if "education_office" in place_types:
        tags.append("Công tác / học tập")

    return " | ".join(unique_keep_order(tags))


def main_purpose_from_tourism_tags(tags):
    """
    Chọn mục đích du lịch chính cho khách sạn.
    """
    text = text_key(tags)

    priority_rules = [
        ("du lich bien", "Du lịch biển / nghỉ dưỡng"),
        ("tham quan van hoa", "Tham quan văn hóa / di tích"),
        ("transit", "Transit / công tác"),
        ("mua sam", "Mua sắm / ăn uống"),
        ("kham pha thanh pho", "Khám phá thành phố"),
        ("thien nhien", "Thiên nhiên / trải nghiệm"),
        ("luu tru y te", "Lưu trú y tế"),
        ("su kien", "Sự kiện / giải trí"),
        ("giao thong", "Di chuyển / trung chuyển"),
        ("cong tac", "Công tác / học tập"),
    ]

    for keyword, purpose in priority_rules:

        if keyword in text:
            return purpose

    return "Lưu trú phổ thông"


def services_from_tourism_tags(tags):
    """
    Gợi ý dịch vụ du lịch theo mục đích.
    """
    text = text_key(tags)

    services = []

    if "transit" in text or "san bay" in text:
        services.extend([
            "Đưa đón sân bay",
            "Check-in nhanh",
            "Gửi hành lý",
            "Taxi/Grab sân bay"
        ])

    if "mua sam" in text or "an uong" in text:
        services.extend([
            "Food tour",
            "Gợi ý quán ăn gần khách sạn",
            "Taxi nội thành",
            "Lịch trình buổi tối"
        ])

    if "du lich bien" in text or "nghi duong" in text:
        services.extend([
            "Tour biển",
            "Thuê xe máy",
            "Gợi ý hải sản",
            "Dịch vụ hồ bơi/spa"
        ])

    if "tham quan van hoa" in text or "di tich" in text:
        services.extend([
            "City tour",
            "Thuê hướng dẫn viên",
            "Thuê xe tham quan",
            "Gợi ý lịch trình văn hóa"
        ])

    if "kham pha thanh pho" in text:
        services.extend([
            "City tour",
            "Gợi ý điểm check-in",
            "Lịch trình 1 ngày quanh khách sạn"
        ])

    if "thien nhien" in text or "trai nghiem" in text:
        services.extend([
            "Tour trải nghiệm",
            "Thuê xe máy",
            "Gợi ý điểm tham quan thiên nhiên"
        ])

    if "luu tru y te" in text:
        services.extend([
            "Phòng yên tĩnh",
            "Dịch vụ giặt ủi",
            "Thang máy",
            "Cửa hàng tiện lợi gần khách sạn"
        ])

    if "su kien" in text:
        services.extend([
            "Taxi sự kiện",
            "Gửi hành lý",
            "Check-out muộn"
        ])

    if "giao thong" in text:
        services.extend([
            "Taxi/Grab",
            "Gợi ý tuyến di chuyển",
            "Gửi hành lý"
        ])

    if "cong tac" in text:
        services.extend([
            "Wifi mạnh",
            "Bàn làm việc",
            "Giặt ủi",
            "Lễ tân 24h"
        ])

    return " | ".join(unique_keep_order(services))


# =====================================================
# HÀM XỬ LÝ TIỆN ÍCH / DỊCH VỤ
# =====================================================

def normalize_facility_group(col_name):
    """
    Chuyển tên cột thành nhóm tiện ích dễ đọc.
    Ví dụ:
    facility_Tiện nghi công cộng -> Tiện nghi công cộng
    general_Tiện ích chung -> Tiện ích chung
    """
    col_name = str(col_name)
    col_name = col_name.replace("facility_", "")
    col_name = col_name.replace("general_", "")
    col_name = normalize_space(col_name)

    if col_name == "":
        return "Tiện ích khác"

    return col_name


def classify_service_suggestion(facility_name):
    """
    Gợi ý nhóm dịch vụ từ tiện ích.
    """
    text = text_key(facility_name)

    if any(k in text for k in ["dua don", "san bay", "taxi", "xe dua don", "cho thue xe", "giu xe", "bai dau xe"]):
        return "Dịch vụ di chuyển"

    if any(k in text for k in ["wifi", "internet", "ban lam viec", "photocopy", "may in", "may chieu", "hoi nghi", "van phong", "phong hop"]):
        return "Dịch vụ công tác"

    if any(k in text for k in ["giat ui", "le tan", "nhan phong", "tra phong", "hanh ly", "bao quan hanh ly", "concierge", "bao ve"]):
        return "Dịch vụ lưu trú"

    if any(k in text for k in ["ho boi", "spa", "massage", "mat xa", "phong gym", "the thao", "xong hoi", "vuon hoa", "karaoke"]):
        return "Dịch vụ nghỉ dưỡng"

    if any(k in text for k in ["nha hang", "bua sang", "quay bar", "am thuc", "cafe", "ca phe", "tiec", "an sang", "an uong"]):
        return "Dịch vụ ăn uống"

    if any(k in text for k in ["tre em", "gia dinh", "thu cung", "cribs", "ghe cho tre", "ho boi tre em"]):
        return "Dịch vụ gia đình"

    if any(k in text for k in ["xe lan", "khuyet tat", "nguoi khuyet tat"]):
        return "Hỗ trợ đặc biệt"

    if any(k in text for k in ["atm", "ngan hang", "cua hang", "sieu thi", "thẩm mỹ", "tham my", "hieu lam toc"]):
        return "Tiện ích lân cận"

    return "Dịch vụ khác"


def merge_dynamic_columns(df, keyword, new_col, split_comma=False):
    """
    Gộp các cột có tên chứa keyword thành 1 cột duy nhất.

    Ví dụ:
    general_Những tiện nghi khác tại Hotel Nikko Saigon
    general_Những tiện nghi khác tại Fusion Original Saigon Centre

    -> general_Những tiện nghi khác
    """
    cols = [
        col for col in df.columns
        if keyword in col
    ]

    if not cols:
        if new_col not in df.columns:
            df[new_col] = ""
        return df

    def merge_row(row):
        items = []

        for col in cols:
            items.extend(
                split_items(
                    row.get(col, ""),
                    split_comma=split_comma
                )
            )

        return " | ".join(
            unique_keep_order(items)
        )

    df[new_col] = df.apply(
        merge_row,
        axis=1
    )

    df = df.drop(
        columns=cols,
        errors="ignore"
    )

    return df.copy()


# =====================================================
# FORMAT EXCEL
# =====================================================

def set_excel_style(writer, sheet_df_map):
    """
    Format Excel cho dễ nhìn.
    Dùng openpyxl.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    normal_alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    for sheet_name, data_df in sheet_df_map.items():

        worksheet = writer.sheets[sheet_name]

        # Freeze dòng tiêu đề
        worksheet.freeze_panes = "A2"

        # Bật filter
        if len(data_df.columns) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        # Style header
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set độ rộng cột
        for col_idx, col_name in enumerate(
            data_df.columns,
            start=1
        ):
            col_letter = worksheet.cell(
                row=1,
                column=col_idx
            ).column_letter

            col_name_str = str(col_name)

            if col_name_str in [
                "hotel_name",
                "hotel_url",
                "address",
                "nearby_places",
                "nearby_places_clean",
                "facilities_clean",
                "service_suggestions",
                "trip_service_suggestions",
                "transaction_items",
                "raw_item",
                "raw_facility_text",
                "raw_nearby_text",
            ]:
                worksheet.column_dimensions[col_letter].width = 45

            elif col_name_str.startswith("facility_"):
                worksheet.column_dimensions[col_letter].width = 35

            elif col_name_str.startswith("general_"):
                worksheet.column_dimensions[col_letter].width = 35

            elif col_name_str.startswith("score_"):
                worksheet.column_dimensions[col_letter].width = 16

            elif col_name_str in [
                "price_clean",
                "review_count_clean",
                "star_rating_clean",
                "overall_rating_clean",
                "distance_m",
            ]:
                worksheet.column_dimensions[col_letter].width = 18

            else:
                worksheet.column_dimensions[col_letter].width = 22

        # Wrap text toàn bộ sheet
        for row in worksheet.iter_rows():

            for cell in row:
                cell.alignment = normal_alignment


# =====================================================
# ĐỌC FILE
# =====================================================

if not INPUT_FILE.exists():
    raise FileNotFoundError(f"Không tìm thấy file input: {INPUT_FILE}")

df = pd.read_excel(INPUT_FILE)

original_row_count = len(df)
original_col_count = len(df.columns)

df.columns = [
    normalize_space(col)
    for col in df.columns
]

print("Số dòng ban đầu:", len(df))
print("Số cột ban đầu:", len(df.columns))


# =====================================================
# 1. XÓA DÒNG TRỐNG + TRÙNG
# =====================================================

df = df.dropna(how="all")

before_duplicate = len(df)

if "hotel_url" in df.columns:
    df = df.drop_duplicates(
        subset=["hotel_url"]
    )
else:
    df = df.drop_duplicates()

removed_duplicate_rows = before_duplicate - len(df)


# =====================================================
# 2. GỘP CÁC CỘT DYNAMIC KHÔNG HỢP LÝ
# =====================================================

df = merge_dynamic_columns(
    df,
    keyword="general_Những tiện nghi khác tại",
    new_col="general_Những tiện nghi khác",
    split_comma=True
)

df = merge_dynamic_columns(
    df,
    keyword="general_Những địa điểm thú vị gần đó tại",
    new_col="general_Những địa điểm thú vị gần đó",
    split_comma=True
)

# Trường hợp cột này đã có sẵn thì giữ nguyên, nếu chưa có thì tạo rỗng
if "general_Những địa điểm thú vị gần đó" not in df.columns:
    df["general_Những địa điểm thú vị gần đó"] = ""


# =====================================================
# 3. XÓA CÁC CỘT KHÔNG HỢP LÝ
# =====================================================

drop_keywords = [
    "general_Số phòng còn trống tai",
    "general_Số phòng còn trống tại",
    "general_Số lầu tại",
]

drop_cols = []

for col in df.columns:

    for keyword in drop_keywords:

        if keyword in col:
            drop_cols.append(col)

drop_cols = list(set(drop_cols))

df = df.drop(
    columns=drop_cols,
    errors="ignore"
)

df = df.copy()

print("Đã xóa số cột rác:", len(drop_cols))


# =====================================================
# 4. LÀM SẠCH CÁC CỘT SỐ
# =====================================================

if "price" in df.columns:
    df["price_clean"] = df["price"].apply(clean_price)
else:
    df["price_clean"] = None

if "star_rating" in df.columns:
    df["star_rating_clean"] = df["star_rating"].apply(clean_float)
else:
    df["star_rating_clean"] = None

if "overall_rating" in df.columns:
    df["overall_rating_clean"] = df["overall_rating"].apply(clean_float)
else:
    df["overall_rating_clean"] = None

if "review_count" in df.columns:
    df["review_count_clean"] = df["review_count"].apply(clean_int)
else:
    df["review_count_clean"] = None


# =====================================================
# 4.1. XÓA DÒNG THIẾU PRICE HOẶC OVERALL_RATING
# =====================================================

before_drop_required = len(df)

df = df.dropna(
    subset=[
        "price_clean",
        "overall_rating_clean"
    ]
).copy()

removed_missing_price_rating = before_drop_required - len(df)

print(
    "Đã xóa số dòng thiếu price hoặc overall_rating:",
    removed_missing_price_rating
)

score_cols = [
    col for col in df.columns
    if col.startswith("score_") and not col.endswith("_clean")
]

for col in score_cols:
    df[col + "_clean"] = df[col].apply(clean_float)


# =====================================================
# 5. TÁCH ĐỊA CHỈ
# =====================================================

if "address" in df.columns:
    df["district"] = df["address"].apply(extract_district)
else:
    df["district"] = ""


# =====================================================
# 6. XỬ LÝ ĐỊA ĐIỂM NỔI BẬT ĐA VÙNG
# =====================================================

nearby_source_cols = []

for col in [
    "nearby_places",
    "general_Điểm đến phổ biến",
    "general_Những địa điểm thú vị gần đó",
]:
    if col in df.columns:
        nearby_source_cols.append(col)


nearby_detail_rows = []

for _, row in df.iterrows():

    hotel_name = row.get("hotel_name", "")
    hotel_url = row.get("hotel_url", "")
    region = row.get("region", "")
    district = row.get("district", "")

    for col in nearby_source_cols:

        # nearby_places dùng dấu |, còn general thường dùng dấu phẩy
        split_comma = col != "nearby_places"

        raw_items = split_items(
            row.get(col, ""),
            split_comma=split_comma
        )

        for raw_item in raw_items:

            place_name, distance_text, distance_m = parse_nearby_item(raw_item)

            if place_name:
                nearby_detail_rows.append({
                    "hotel_name": hotel_name,
                    "hotel_url": hotel_url,
                    "region": region,
                    "district": district,
                    "source_col": col,
                    "raw_item": raw_item,
                    "place_name": place_name,
                    "distance_text": distance_text,
                    "distance_m": distance_m,
                    "place_type": classify_place_type(
                        place_name,
                        region
                    ),
                    "distance_level": distance_level(distance_m),
                })


nearby_detail_df = pd.DataFrame(
    nearby_detail_rows,
    columns=[
        "hotel_name",
        "hotel_url",
        "region",
        "district",
        "source_col",
        "raw_item",
        "place_name",
        "distance_text",
        "distance_m",
        "place_type",
        "distance_level",
    ]
)

if len(nearby_detail_df) > 0:
    nearby_detail_df = nearby_detail_df.drop_duplicates()


def get_nearby_for_hotel(row):
    hotel_url = row.get("hotel_url", "")
    hotel_name = row.get("hotel_name", "")

    if "hotel_url" in nearby_detail_df.columns and hotel_url:
        temp = nearby_detail_df[
            nearby_detail_df["hotel_url"] == hotel_url
        ]
    else:
        temp = nearby_detail_df[
            nearby_detail_df["hotel_name"] == hotel_name
        ]

    return temp


def build_nearby_places_from_detail(row):
    temp = get_nearby_for_hotel(row)

    places = temp["place_name"].tolist()

    return " | ".join(unique_keep_order(places))


def build_tourism_tags_from_detail(row):
    temp = get_nearby_for_hotel(row)

    place_types = temp["place_type"].tolist()

    return tourism_tags_from_place_types(place_types)


def build_closest_place_info(row, place_type):
    temp = get_nearby_for_hotel(row)

    temp = temp[
        (temp["place_type"] == place_type)
        & (temp["distance_m"].notna())
    ]

    if len(temp) == 0:
        return ""

    temp = temp.sort_values("distance_m")
    first = temp.iloc[0]

    return f"{first['place_name']} ({first['distance_text']})"


def build_min_distance(row, place_type):
    temp = get_nearby_for_hotel(row)

    temp = temp[
        (temp["place_type"] == place_type)
        & (temp["distance_m"].notna())
    ]

    if len(temp) == 0:
        return None

    return int(temp["distance_m"].min())


def count_place_type(row, place_type):
    temp = get_nearby_for_hotel(row)

    temp = temp[
        temp["place_type"] == place_type
    ]

    return len(temp)


def count_walkable_places(row):
    temp = get_nearby_for_hotel(row)

    temp = temp[
        (temp["distance_m"].notna())
        & (temp["distance_m"] <= 1000)
    ]

    return len(temp)


df["nearby_places_clean"] = df.apply(
    build_nearby_places_from_detail,
    axis=1
)

df["tourism_tags"] = df.apply(
    build_tourism_tags_from_detail,
    axis=1
)

df["main_tourism_purpose"] = df["tourism_tags"].apply(
    main_purpose_from_tourism_tags
)

df["trip_service_suggestions"] = df["tourism_tags"].apply(
    services_from_tourism_tags
)

# Cột closest theo từng nhóm địa điểm để web dễ lọc/gợi ý
df["closest_airport"] = df.apply(
    lambda row: build_closest_place_info(row, "airport"),
    axis=1
)

df["closest_beach"] = df.apply(
    lambda row: build_closest_place_info(row, "beach_relax"),
    axis=1
)

df["closest_shopping_food"] = df.apply(
    lambda row: build_closest_place_info(row, "shopping_food"),
    axis=1
)

df["closest_heritage_culture"] = df.apply(
    lambda row: build_closest_place_info(row, "heritage_culture"),
    axis=1
)

df["closest_city_sightseeing"] = df.apply(
    lambda row: build_closest_place_info(row, "city_sightseeing"),
    axis=1
)

df["closest_nature_experience"] = df.apply(
    lambda row: build_closest_place_info(row, "nature_experience"),
    axis=1
)

df["closest_medical"] = df.apply(
    lambda row: build_closest_place_info(row, "medical"),
    axis=1
)

df["closest_transport"] = df.apply(
    lambda row: build_closest_place_info(row, "transport"),
    axis=1
)

# Cột khoảng cách min theo mét
df["closest_airport_m"] = df.apply(
    lambda row: build_min_distance(row, "airport"),
    axis=1
)

df["closest_beach_m"] = df.apply(
    lambda row: build_min_distance(row, "beach_relax"),
    axis=1
)

df["closest_shopping_food_m"] = df.apply(
    lambda row: build_min_distance(row, "shopping_food"),
    axis=1
)

df["closest_heritage_culture_m"] = df.apply(
    lambda row: build_min_distance(row, "heritage_culture"),
    axis=1
)

df["closest_city_sightseeing_m"] = df.apply(
    lambda row: build_min_distance(row, "city_sightseeing"),
    axis=1
)

df["closest_nature_experience_m"] = df.apply(
    lambda row: build_min_distance(row, "nature_experience"),
    axis=1
)

df["closest_medical_m"] = df.apply(
    lambda row: build_min_distance(row, "medical"),
    axis=1
)

df["closest_transport_m"] = df.apply(
    lambda row: build_min_distance(row, "transport"),
    axis=1
)

# Count theo nhóm địa điểm
df["walkable_place_count"] = df.apply(
    count_walkable_places,
    axis=1
)

for place_type in PLACE_TYPE_RULES.keys():
    df[f"nearby_count_{place_type}"] = df.apply(
        lambda row, pt=place_type: count_place_type(row, pt),
        axis=1
    )


# =====================================================
# 7. XỬ LÝ TIỆN ÍCH CHI TIẾT, KHÔNG MẤT DỮ LIỆU
# =====================================================

facility_cols = [
    col for col in df.columns
    if col.startswith("facility_")
]

extra_facility_cols = []

for col in [
    "general_Tiện ích chung",
    "general_Những tiện nghi khác",
    "general_Có ăn sáng",
]:
    if col in df.columns:
        extra_facility_cols.append(col)

facility_source_cols = facility_cols + extra_facility_cols

facility_detail_rows = []

for _, row in df.iterrows():

    hotel_name = row.get("hotel_name", "")
    hotel_url = row.get("hotel_url", "")
    region = row.get("region", "")
    district = row.get("district", "")

    for col in facility_source_cols:

        # general_Tiện ích chung và general_Những tiện nghi khác thường ngăn cách bằng dấu phẩy
        split_comma = col in extra_facility_cols

        raw_items = split_items(
            row.get(col, ""),
            split_comma=split_comma
        )

        facility_group = normalize_facility_group(col)

        for facility_name in raw_items:

            if facility_name:
                facility_detail_rows.append({
                    "hotel_name": hotel_name,
                    "hotel_url": hotel_url,
                    "region": region,
                    "district": district,
                    "source_col": col,
                    "facility_group": facility_group,
                    "facility_name": facility_name,
                    "service_suggestion": classify_service_suggestion(facility_name),
                })


facility_detail_df = pd.DataFrame(
    facility_detail_rows,
    columns=[
        "hotel_name",
        "hotel_url",
        "region",
        "district",
        "source_col",
        "facility_group",
        "facility_name",
        "service_suggestion",
    ]
)

if len(facility_detail_df) > 0:
    facility_detail_df = facility_detail_df.drop_duplicates()


def get_facilities_for_hotel(row):
    hotel_url = row.get("hotel_url", "")
    hotel_name = row.get("hotel_name", "")

    if "hotel_url" in facility_detail_df.columns and hotel_url:
        temp = facility_detail_df[
            facility_detail_df["hotel_url"] == hotel_url
        ]
    else:
        temp = facility_detail_df[
            facility_detail_df["hotel_name"] == hotel_name
        ]

    return temp


def build_facilities_from_detail(row):
    temp = get_facilities_for_hotel(row)

    facilities = temp["facility_name"].tolist()

    return " | ".join(unique_keep_order(facilities))


def build_facility_groups_from_detail(row):
    temp = get_facilities_for_hotel(row)

    groups = temp["facility_group"].tolist()

    return " | ".join(unique_keep_order(groups))


def build_services_from_detail(row):
    temp = get_facilities_for_hotel(row)

    services = temp["service_suggestion"].tolist()

    return " | ".join(unique_keep_order(services))


df["facilities_clean"] = df.apply(
    build_facilities_from_detail,
    axis=1
)

df["facility_groups_clean"] = df.apply(
    build_facility_groups_from_detail,
    axis=1
)

df["service_suggestions"] = df.apply(
    build_services_from_detail,
    axis=1
)

# Gộp dịch vụ từ tiện ích + dịch vụ theo mục đích chuyến đi

def build_all_service_suggestions(row):
    items = []
    items.extend(split_items(row.get("service_suggestions", "")))
    items.extend(split_items(row.get("trip_service_suggestions", "")))
    return " | ".join(unique_keep_order(items))


df["all_service_suggestions"] = df.apply(
    build_all_service_suggestions,
    axis=1
)


# =====================================================
# 8. TẠO CỘT PHÂN NHÓM
# =====================================================

if "price_clean" in df.columns:
    df["price_level"] = df["price_clean"].apply(price_level)
else:
    df["price_level"] = ""

if "overall_rating_clean" in df.columns:
    df["rating_level"] = df["overall_rating_clean"].apply(rating_level)
else:
    df["rating_level"] = ""


# =====================================================
# 9. TẠO TRANSACTION CHO APRIORI
# =====================================================

def build_transaction(row):
    items = []

    hotel_type_cols = [
        "region",
        "price_level",
        "rating_level",
        "district",
        "tourism_tags",
        "main_tourism_purpose",
        "all_service_suggestions",
    ]

    for col in hotel_type_cols:

        values = split_items(
            row.get(col, ""),
            split_comma=False
        )

        for value in values:
            prefix = col.replace("_level", "")
            item = make_item(
                prefix,
                value
            )

            if item:
                items.append(item)

    if not pd.isna(
        row.get("star_rating_clean", None)
    ):
        star = int(row["star_rating_clean"])
        items.append(f"star_{star}_sao")

    facility_values = split_items(
        row.get("facilities_clean", ""),
        split_comma=False
    )

    for facility in facility_values:
        item = make_item(
            "facility",
            facility
        )

        if item:
            items.append(item)

    facility_group_values = split_items(
        row.get("facility_groups_clean", ""),
        split_comma=False
    )

    for group in facility_group_values:
        item = make_item(
            "facility_group",
            group
        )

        if item:
            items.append(item)

    # Thêm nhóm địa điểm vào transaction, không thêm tên địa điểm thật để tránh quá thưa dữ liệu
    for place_type in PLACE_TYPE_RULES.keys():
        count_col = f"nearby_count_{place_type}"

        if row.get(count_col, 0) and row.get(count_col, 0) > 0:
            items.append(f"place_type_{place_type}")

    if row.get("walkable_place_count", 0) and row.get("walkable_place_count", 0) > 0:
        items.append("distance_co_dia_diem_di_bo_duoc")

    return " | ".join(
        unique_keep_order(items)
    )


df["transaction_items"] = df.apply(
    build_transaction,
    axis=1
)


# =====================================================
# 10. TẠO FILE ONE-HOT CHO APRIORI
# =====================================================

transactions = df["transaction_items"].apply(
    lambda x: split_items(
        x,
        split_comma=False
    )
)

all_items = sorted(
    set(
        item
        for trans in transactions
        for item in trans
    )
)

onehot_rows = []

for trans in transactions:
    trans_set = set(trans)

    row = {
        item: int(item in trans_set)
        for item in all_items
    }

    onehot_rows.append(row)

onehot_df = pd.DataFrame(onehot_rows)

if "hotel_name" in df.columns:
    onehot_df.insert(
        0,
        "hotel_name",
        df["hotel_name"].values
    )


# =====================================================
# 11. TẠO CÁC SHEET DỄ NHÌN
# =====================================================

nearby_df = nearby_detail_df.copy()
facility_df = facility_detail_df.copy()

transaction_df = df[
    [
        "hotel_name",
        "hotel_url",
        "transaction_items"
    ]
].copy()

service_rows = []

for _, row in df.iterrows():
    hotel_name = row.get("hotel_name", "")
    hotel_url = row.get("hotel_url", "")
    region = row.get("region", "")
    district = row.get("district", "")

    services = split_items(
        row.get("all_service_suggestions", ""),
        split_comma=False
    )

    for service in services:
        service_rows.append({
            "hotel_name": hotel_name,
            "hotel_url": hotel_url,
            "region": region,
            "district": district,
            "service_suggestion": service,
            "main_tourism_purpose": row.get("main_tourism_purpose", ""),
        })

service_df = pd.DataFrame(
    service_rows,
    columns=[
        "hotel_name",
        "hotel_url",
        "region",
        "district",
        "service_suggestion",
        "main_tourism_purpose",
    ]
)

summary_df = pd.DataFrame({
    "Chỉ số": [
        "Số khách sạn ban đầu",
        "Số khách sạn sau xử lý",
        "Số cột ban đầu",
        "Số cột sau xử lý",
        "Số dòng trùng đã xóa",
        "Số dòng thiếu price/rating đã xóa",
        "Số cột rác đã xóa",
        "Số địa điểm nearby đã tách",
        "Số tiện ích đã tách",
        "Số dịch vụ gợi ý đã tách",
        "Số item Apriori",
    ],
    "Giá trị": [
        original_row_count,
        len(df),
        original_col_count,
        len(df.columns),
        removed_duplicate_rows,
        removed_missing_price_rating,
        len(drop_cols),
        len(nearby_df),
        len(facility_df),
        len(service_df),
        len(onehot_df.columns) - 1
        if "hotel_name" in onehot_df.columns
        else len(onehot_df.columns),
    ]
})


# =====================================================
# 12. LƯU FILE
# =====================================================

df.to_excel(
    CLEANED_FILE,
    index=False
)

transaction_df.to_csv(
    TRANSACTION_FILE,
    index=False,
    encoding="utf-8-sig"
)

onehot_df.to_csv(
    ONEHOT_FILE,
    index=False,
    encoding="utf-8-sig"
)

sheet_df_map = {
    "summary": summary_df,
    "cleaned_data": df,
    "nearby_places": nearby_df,
    "facilities": facility_df,
    "services": service_df,
    "transactions": transaction_df,
    "apriori_onehot": onehot_df,
}

with pd.ExcelWriter(
    REPORT_FILE,
    engine="openpyxl"
) as writer:

    for sheet_name, sheet_df in sheet_df_map.items():
        sheet_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )

    set_excel_style(
        writer,
        sheet_df_map
    )


print("Số dòng sau xử lý:", len(df))
print("Số cột sau xử lý:", len(df.columns))
print("Số địa điểm đã tách:", len(nearby_df))
print("Số tiện ích đã tách:", len(facility_df))

print("Đã lưu file cleaned:", CLEANED_FILE)
print("Đã lưu file transaction:", TRANSACTION_FILE)
print("Đã lưu file one-hot:", ONEHOT_FILE)
print("Đã lưu file Excel report:", REPORT_FILE)
